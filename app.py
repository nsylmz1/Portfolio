from flask import Flask, request, jsonify, send_file
import firebase_admin
from firebase_admin import credentials, firestore
import io
from openpyxl import Workbook
from datetime import datetime, timedelta
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from flask_cors import CORS
from firebase_admin.firestore import FieldFilter
import openpyxl.styles

# Firebase yapılandırması
cred = credentials.Certificate("egorevtakip-firebase-adminsdk-fbsvc-c219d7211d.json")
firebase_admin.initialize_app(cred, {
    'projectId': 'gorev-takip-cc22e',
    'databaseURL': 'https://gorev-takip-cc22e.firebaseio.com',
    'storageBucket': 'gorev-takip-cc22e.firebasestorage.app',
    'httpTimeout': 30  # 30 saniye timeout
})
db = firestore.client()

# Flask uygulaması
app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*", "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"], "allow_headers": ["Content-Type", "Authorization", "X-User-Email"]}})
 
# E-posta bildirim ayarları
ADMIN_EMAIL = "xxenesyilmazxx@gmail.com"  # Admin e-posta adresini güncelledim
APP_PASSWORD = "katv ttar rxtp iurw"  # E-posta şifresi

def send_task_notification(recipient_email, task_title, task_description, due_date):
    """Kullanıcıya yeni görev bildirimi gönderir"""
    try:
        # E-posta içeriği oluştur
        message = MIMEMultipart()
        message['From'] = ADMIN_EMAIL
        message['To'] = recipient_email
        message['Subject'] = "Yeni Görev Ataması: " + task_title
        
        # E-posta gövdesi
        body = f"""Merhaba,

Size yeni bir görev atandı:

Başlık: {task_title}
Açıklama: {task_description}
Bitirilmesi İstenen Tarih: {due_date}

Görev takip sistemine giriş yaparak görevi görüntüleyebilirsiniz.

İyi çalışmalar,
Görev Takip Sistemi
"""
        message.attach(MIMEText(body, 'plain'))
        
        # E-posta gönderme
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(ADMIN_EMAIL, APP_PASSWORD)
        server.send_message(message)
        server.quit()
        
        print(f"E-posta başarıyla gönderildi: {recipient_email}")
        return True
    except Exception as e:
        print(f"E-posta gönderme hatası: {e}")
        # E-posta gönderimi başarısız olsa bile görev ekleme işlemine devam et
        return False

@app.route('/')
def home():
    return "Merhaba, Görev Takip Uygulamasına Hoşgeldiniz!"

@app.route('/gorev', methods=['POST'])
def create_task():
    try:
        data = request.get_json()
        user_email = data.get('user_email')
        
        # Kullanıcıyı bul
        users_ref = db.collection('users')
        user_query = users_ref.where(filter=FieldFilter("email", "==", user_email)).limit(1).get()
        
        if not user_query:
            return jsonify({'error': 'Kullanıcı bulunamadı'}), 404
            
        user_doc = user_query[0]
        user_data = user_doc.to_dict()
        
        # Admin kullanıcıların görev eklemesini engelle
        if user_data.get('role') == 'admin':
            return jsonify({'error': 'Admin kullanıcılar görev ekleyemez'}), 403
        
        # Görev verilerini hazırla
        task_data = {
            'title': data.get('title'),
            'description': data.get('description'),
            'due_date': data.get('due_date'),
            'status': data.get('status', 'pending'),
            'assigned_to': data.get('assigned_to'),
            'assigned_by': user_email,
            'assigned_date': datetime.now().strftime('%Y-%m-%d'),
            'user_id': user_doc.id,
            'user_email': user_email
        }
        
        # Görevi oluştur
        task_ref = db.collection('tasks').document()
        task_ref.set(task_data)
        
        # Bildirim oluştur
        notification_data = {
            'title': 'Yeni Görev',
            'message': f'Size yeni bir görev atandı: {task_data["title"]}',
            'email': task_data['assigned_to'],  # user_email yerine email kullanıyoruz
            'task_id': task_ref.id,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'is_read': False
        }
        
        db.collection('notifications').add(notification_data)
        
        # E-posta gönder
        try:
            send_task_notification(
                task_data['assigned_to'],
                task_data['title'],
                task_data['description'],
                task_data['due_date']
            )
        except Exception as e:
            print(f"E-posta gönderme hatası: {str(e)}")
        
        return jsonify({'message': 'Görev başarıyla oluşturuldu', 'task_id': task_ref.id}), 201
        
    except Exception as e:
        print(f"Görev oluşturma hatası: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/gorevler', methods=['GET'])
def tum_gorevler():
    try:
        # Filtreleme için tarih parametresi
        filter_date = request.args.get('date')
        
        if filter_date:
            # Belirli bir tarihe göre filtrele (atanma tarihi veya bitiş tarihi)
            filter_type = request.args.get('filter_type', 'due_date')  # Varsayılan olarak bitiş tarihine göre
            
            gorevler_ref = db.collection("tasks").where(filter_type, "==", filter_date)
        else:
            # Filtre yoksa tüm görevleri getir
            gorevler_ref = db.collection("tasks")
            
        docs = gorevler_ref.stream()
        gorevler = []
        
        for doc in docs:
            gorev = doc.to_dict()
            gorev['id'] = doc.id
            
            # Durum Türkçeleştirme
            gorev['durum_tr'] = get_status_turkish(gorev.get('status', 'unknown'))
            
            gorevler.append(gorev)
            
        return jsonify(gorevler), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/gorevler/<string:user_email>', methods=['GET'])
def gorevleri_getir(user_email):
    try:
        # Filtreleme için tarih parametresi
        filter_date = request.args.get('date')
        
        if filter_date:
            # Belirli bir tarihe ve kullanıcıya göre filtrele
            filter_type = request.args.get('filter_type', 'due_date')
            gorevler_ref = db.collection("tasks").where("assigned_to", "==", user_email).where(filter_type, "==", filter_date)
        else:
            # Sadece kullanıcıya göre filtrele
            gorevler_ref = db.collection("tasks").where("assigned_to", "==", user_email)
            
        docs = gorevler_ref.stream()
        gorevler = []
        
        for doc in docs:
            gorev = doc.to_dict()
            gorev['id'] = doc.id
            
            # Durum Türkçeleştirme
            gorev['durum_tr'] = get_status_turkish(gorev.get('status', 'unknown'))
            
            gorevler.append(gorev)
            
        return jsonify(gorevler), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/gorev/<task_id>', methods=['GET', 'PUT'])
def update_task(task_id):
    if request.method == 'GET':
        try:
            # Görevi bul
            task_ref = db.collection('tasks').document(task_id)
            task = task_ref.get()
            
            if not task.exists:
                return jsonify({'error': 'Görev bulunamadı'}), 404
                
            task_data = task.to_dict()
            task_data['id'] = task.id
            return jsonify(task_data), 200
            
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    elif request.method == 'PUT':
        try:
            data = request.json
            
            # Görevi bul
            task_ref = db.collection('tasks').document(task_id)
            task = task_ref.get()
            
            if not task.exists:
                return jsonify({'error': 'Görev bulunamadı'}), 404
            
            # Görevi güncelle
            task_ref.update(data)
            
            return jsonify({'message': 'Görev başarıyla güncellendi'}), 200
            
        except Exception as e:
            return jsonify({'error': str(e)}), 500

@app.route('/gorev/<task_id>', methods=['DELETE'])
def delete_task(task_id):
    try:
        # Görevi bul
        task_ref = db.collection('tasks').document(task_id)
        task = task_ref.get()
        
        if not task.exists:
            return jsonify({'error': 'Görev bulunamadı'}), 404
        
        # Görevi sil
        task_ref.delete()
        
        return jsonify({'message': 'Görev başarıyla silindi'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/kullanicilar', methods=['GET'])
def kullanicilari_getir():
    try:
        kullanicilar_ref = db.collection("users")
        docs = kullanicilar_ref.stream()
        kullanicilar = []
        for doc in docs:
            kullanici = doc.to_dict()
            kullanici['id'] = doc.id
            kullanicilar.append(kullanici)
        return jsonify(kullanicilar), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def safe_get_value(data, key, default="Belirtilmemiş"):
    """Güvenli veri alma fonksiyonu"""
    value = data.get(key)
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return default
    return str(value).strip()

@app.route('/rapor', methods=['GET'])
def rapor_indir():
    try:
        # Kullanıcı rolünü kontrol et
        user_email = request.headers.get('X-User-Email')
        if not user_email:
            return jsonify({"error": "Kullanıcı bilgisi bulunamadı"}), 401

        user_ref = db.collection("users").where("email", "==", user_email).limit(1).get()
        if not user_ref:
            return jsonify({"error": "Kullanıcı bulunamadı"}), 404

        user_data = user_ref[0].to_dict()
        if user_data.get('role') != 'superAdmin':
            return jsonify({"error": "Bu işlem için yetkiniz yok"}), 403

        # Filtreleme için tarih parametresi
        filter_date = request.args.get('date', datetime.now().strftime("%Y-%m-%d"))
        filter_type = request.args.get('filter_type', 'due_date')  # Varsayılan olarak bitiş tarihine göre
        
        # Belirtilen tarihe göre görevleri filtrele
        gorevler_ref = db.collection("tasks").where(filter_type, "==", filter_date)
        docs = gorevler_ref.stream()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Görev Raporu"
        
        # Başlıkları ayarla
        headers = ["Başlık", "Açıklama", "Durum", "Atanan", "Atanma Tarihi", "Bitirilmesi İstenen Tarih"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = cell.font.copy(bold=True)
        
        # Görevleri ekle
        row = 2
        for doc in docs:
            g = doc.to_dict()
            
            # Güvenli veri alma
            title = safe_get_value(g, "title", "Başlık Belirtilmemiş")
            description = safe_get_value(g, "description", "Açıklama Belirtilmemiş") 
            status = g.get("status", "pending")
            status_tr = get_status_turkish(status)
            assigned_to = safe_get_value(g, "assigned_to", "Atanmamış")
            assigned_date = safe_get_value(g, "assigned_date", "Tarih Belirtilmemiş")
            due_date = safe_get_value(g, "due_date", "Tarih Belirtilmemiş")
            
            # Hücrelere değerleri ata
            ws.cell(row=row, column=1, value=title)
            ws.cell(row=row, column=2, value=description)
            ws.cell(row=row, column=3, value=status_tr)
            ws.cell(row=row, column=4, value=assigned_to)
            ws.cell(row=row, column=5, value=assigned_date)
            ws.cell(row=row, column=6, value=due_date)
            
            # Hücre stillerini ayarla
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='center')
            
            row += 1
        
        # Sütun genişliklerini ayarla
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Başlık satırını biçimlendir
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Dosya adına tarih bilgisi ekle
        filename = f"gorev_raporu_{filter_date}_{filter_type}.xlsx"
        
        response = send_file(
            output,
            download_name=filename,
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        # Tarayıcıların dosyayı doğru şekilde işlemesi için ek başlıklar
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        return response
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Durum değerlerini Türkçeleştir
def get_status_turkish(status):
    status_dict = {
        'pending': 'Beklemede',
        'in-progress': 'Devam Ediyor',
        'completed': 'Tamamlandı',
        'unknown': 'Bilinmiyor'
    }
    return status_dict.get(status, 'Bilinmiyor')

def get_user_role():
    try:
        user_email = request.headers.get('X-User-Email')
        if not user_email:
            print("Kullanıcı e-postası bulunamadı")
            return None

        print(f"Kullanıcı rolü kontrol ediliyor: {user_email}")
        user_ref = db.collection("users").where("email", "==", user_email).limit(1).get()
        
        if not user_ref:
            print(f"Kullanıcı bulunamadı: {user_email}")
            return None

        user_data = user_ref[0].to_dict()
        role = user_data.get('role')
        print(f"Kullanıcı rolü: {role}")
        return role
    except Exception as e:
        print(f"Rol kontrolü hatası: {e}")
        return None

@app.route('/kullanici/<string:user_id>', methods=['GET'])
def get_user_role(user_id):
    try:
        user_ref = db.collection("users").document(user_id)
        user_doc = user_ref.get()
        if user_doc.exists:
            return jsonify(user_doc.to_dict()), 200
        return jsonify({"error": "Kullanıcı bulunamadı"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/kullanici/<string:user_id>', methods=['PUT'])
def kullanici_guncelle(user_id):
    try:
        data = request.json
        if not data:
            return jsonify({"error": "Güncelleme bilgisi yok"}), 400
        
        user_ref = db.collection("users").document(user_id)
        user_ref.update(data)
        return jsonify({"message": "Kullanıcı güncellendi"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/kullanici-rol/<user_id>', methods=['PUT'])
def kullanici_rol_guncelle(user_id):
    try:
        # Kullanıcı rolünü kontrol et
        user_role = get_user_role()
        if not user_role or user_role not in ['superAdmin', 'admin']:
            return jsonify({'error': 'Yetkisiz erişim'}), 401

        data = request.get_json()
        new_role = data.get('role')

        # Geçerli roller
        valid_roles = ['user', 'admin', 'superAdmin']

        # Admin için SuperAdmin rolü atama kontrolü
        if user_role == 'admin' and new_role == 'superAdmin':
            return jsonify({'error': 'Admin kullanıcıları SuperAdmin rolü atayamaz'}), 403

        # Geçerli rol kontrolü
        if new_role not in valid_roles:
            return jsonify({'error': 'Geçersiz rol'}), 400

        # Kullanıcıyı güncelle
        user_ref = db.collection('users').document(user_id)
        user_ref.update({'role': new_role})

        return jsonify({'message': 'Kullanıcı rolü güncellendi'}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Bildirimleri getir
@app.route('/bildirimler/<email>', methods=['GET'])
def get_notifications(email):
    try:
        # Kullanıcının bildirimlerini getir
        notifications_ref = db.collection('notifications')
        notifications = notifications_ref.where(filter=FieldFilter("email", "==", email)).order_by("created_at", direction=firestore.Query.DESCENDING).stream()
        
        notifications_list = []
        for notification in notifications:
            notification_data = notification.to_dict()
            notification_data['id'] = notification.id
            notifications_list.append(notification_data)
            
        return jsonify(notifications_list), 200
    except Exception as e:
        print(f"Bildirim getirme hatası: {str(e)}")
        return jsonify({"error": "Bildirimler getirilirken bir hata oluştu"}), 500

# Görevleri getir
@app.route('/gorevler', methods=['GET'])
def get_tasks():
    try:
        tasks_ref = db.collection('tasks')
        tasks = tasks_ref.order_by("created_at", direction=firestore.Query.DESCENDING).stream()
        
        tasks_list = []
        for task in tasks:
            task_data = task.to_dict()
            task_data['id'] = task.id
            tasks_list.append(task_data)
            
        return jsonify(tasks_list)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Kullanıcının görevlerini getir
@app.route('/gorevler/<email>', methods=['GET'])
def get_user_tasks(email):
    try:
        # Kullanıcıyı bul
        user_ref = db.collection('users').where(filter=FieldFilter("email", "==", email)).limit(1)
        user_docs = user_ref.stream()
        user_doc = next(user_docs, None)
        
        if not user_doc:
            return jsonify({"error": "Kullanıcı bulunamadı"}), 404
            
        user_data = user_doc.to_dict()
        user_id = user_doc.id
        
        # Kullanıcının görevlerini getir
        tasks_ref = db.collection('tasks')
        tasks = tasks_ref.where(filter=FieldFilter("assigned_to", "==", user_id)).order_by("created_at", direction=firestore.Query.DESCENDING).stream()
        
        tasks_list = []
        for task in tasks:
            task_data = task.to_dict()
            task_data['id'] = task.id
            tasks_list.append(task_data)
            
        return jsonify(tasks_list)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Takvim görevlerini getir
@app.route('/takvim-gorevler', methods=['GET'])
def get_calendar_tasks():
    try:
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        
        if not year or not month:
            return jsonify({"error": "Yıl ve ay parametreleri gerekli"}), 400
            
        # Ayın başlangıç ve bitiş tarihlerini hesapla
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1) - timedelta(days=1)
            
        # Görevleri getir
        gorevler_ref = db.collection('tasks')
        gorevler = gorevler_ref.where(filter=FieldFilter("due_date", ">=", start_date.strftime('%Y-%m-%d'))).where(filter=FieldFilter("due_date", "<=", end_date.strftime('%Y-%m-%d'))).stream()
        
        gorevler_list = []
        for gorev in gorevler:
            gorev_data = gorev.to_dict()
            gorev_data['id'] = gorev.id
            gorevler_list.append(gorev_data)
            
        return jsonify(gorevler_list)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Bildirimi okundu olarak işaretle
@app.route('/bildirim-okundu/<notification_id>', methods=['PUT'])
def mark_notification_read(notification_id):
    try:
        notification_ref = db.collection('notifications').document(notification_id)
        notification_ref.update({'read': True})
        return jsonify({"message": "Bildirim okundu olarak işaretlendi"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Bildirim oluştur
@app.route('/bildirim', methods=['POST'])
def create_notification():
    try:
        data = request.get_json()
        
        # Bildirim verilerini hazırla
        notification_data = {
            'title': data.get('title'),
            'message': data.get('message'),
            'email': data.get('user_email'),
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'is_read': False
        }
        
        # Bildirimi oluştur
        db.collection('notifications').add(notification_data)
        
        return jsonify({'message': 'Bildirim başarıyla oluşturuldu'}), 201
        
    except Exception as e:
        print(f"Bildirim oluşturma hatası: {str(e)}")
        return jsonify({'error': str(e)}), 500

# Görev durumu güncelleme endpoint'ini güncelle
@app.route('/gorev-durum/<task_id>', methods=['PUT', 'OPTIONS'])
def update_task_status(task_id):
    if request.method == 'OPTIONS':
        return '', 200
        
    try:
        data = request.get_json()
        new_status = data.get('status')
        
        if not new_status:
            return jsonify({'error': 'Durum bilgisi gerekli'}), 400
            
        # Eğer durum "in-progress" ise ve kullanıcı görevin sahibiyse, adım ekleme sayfasına yönlendir
        if new_status == 'in-progress':
            # Bu durumda frontend'de adım ekleme modalı açılacak
            pass
            
        # Görevi bul
        task_ref = db.collection('tasks').document(task_id)
        task = task_ref.get()
        
        if not task.exists:
            return jsonify({'error': 'Görev bulunamadı'}), 404
            
        # Görev durumunu güncelle
        task_ref.update({
            'status': new_status,
            'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
        # Bildirim oluştur
        task_data = task.to_dict()
        notification_data = {
            'title': 'Görev Durumu Güncellendi',
            'message': f'"{task_data["title"]}" görevinin durumu {get_status_turkish(new_status)} olarak güncellendi',
            'email': task_data['assigned_to'],
            'task_id': task_id,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'is_read': False
        }
        
        db.collection('notifications').add(notification_data)
        
        return jsonify({'message': 'Görev durumu başarıyla güncellendi'}), 200
        
    except Exception as e:
        print(f"Görev durumu güncelleme hatası: {str(e)}")
        return jsonify({'error': str(e)}), 500

# Göreve adım ekleme
@app.route('/gorev/<task_id>/adim', methods=['POST'])
def add_task_step(task_id):
    try:
        data = request.get_json()
        user_email = request.headers.get('X-User-Email')
        
        if not user_email:
            return jsonify({'error': 'Kullanıcı bilgisi gerekli'}), 401
            
        # Görevi kontrol et
        task_ref = db.collection('tasks').document(task_id)
        task = task_ref.get()
        
        if not task.exists:
            return jsonify({'error': 'Görev bulunamadı'}), 404
            
        task_data = task.to_dict()
        
        # Sadece görevin sahibi adım ekleyebilir
        if task_data.get('assigned_to') != user_email:
            return jsonify({'error': 'Bu göreve sadece atanan kişi adım ekleyebilir'}), 403
            
        # Adım verilerini hazırla
        step_data = {
            'task_id': task_id,
            'title': data.get('title'),
            'description': data.get('description', ''),
            'status': data.get('status', 'pending'),
            'created_by': user_email,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        # Adımı kaydet
        step_ref = db.collection('task_steps').document()
        step_ref.set(step_data)
        
        # Görevin durumunu "in-progress" yap
        task_ref.update({
            'status': 'in-progress',
            'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
        return jsonify({'message': 'Adım başarıyla eklendi', 'step_id': step_ref.id}), 201
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Görevin adımlarını getir
@app.route('/gorev/<task_id>/adimlar', methods=['GET'])
def get_task_steps(task_id):
    try:
        user_email = request.headers.get('X-User-Email')
        
        if not user_email:
            return jsonify({'error': 'Kullanıcı bilgisi gerekli'}), 401
            
        # Kullanıcı rolünü kontrol et
        user_ref = db.collection("users").where("email", "==", user_email).limit(1).get()
        if not user_ref:
            return jsonify({'error': 'Kullanıcı bulunamadı'}), 404
            
        user_data = user_ref[0].to_dict()
        user_role = user_data.get('role')
        
        # Görevin adımlarını getir
        steps_ref = db.collection('task_steps').where('task_id', '==', task_id).order_by('created_at', direction=firestore.Query.ASCENDING)
        steps = steps_ref.stream()
        
        steps_list = []
        for step in steps:
            step_data = step.to_dict()
            step_data['id'] = step.id
            step_data['status_tr'] = get_status_turkish(step_data.get('status', 'pending'))
            steps_list.append(step_data)
            
        # Görev bilgisini de ekle
        task_ref = db.collection('tasks').document(task_id)
        task = task_ref.get()
        
        if task.exists:
            task_data = task.to_dict()
            # Sadece admin/superAdmin veya görevin sahibi görebilir
            if user_role in ['admin', 'superAdmin'] or task_data.get('assigned_to') == user_email:
                return jsonify({
                    'task': task_data,
                    'steps': steps_list
                }), 200
            else:
                return jsonify({'error': 'Bu görevin adımlarını görme yetkiniz yok'}), 403
        else:
            return jsonify({'error': 'Görev bulunamadı'}), 404
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Adım güncelleme
@app.route('/adim/<step_id>', methods=['PUT'])
def update_task_step(step_id):
    try:
        data = request.get_json()
        user_email = request.headers.get('X-User-Email')
        
        if not user_email:
            return jsonify({'error': 'Kullanıcı bilgisi gerekli'}), 401
            
        # Adımı bul
        step_ref = db.collection('task_steps').document(step_id)
        step = step_ref.get()
        
        if not step.exists:
            return jsonify({'error': 'Adım bulunamadı'}), 404
            
        step_data = step.to_dict()
        
        # Sadece adımı oluşturan kişi güncelleyebilir
        if step_data.get('created_by') != user_email:
            return jsonify({'error': 'Bu adımı sadece oluşturan kişi güncelleyebilir'}), 403
            
        # Güncelleme verilerini hazırla
        update_data = {
            'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        # Hangi alanlar güncellenmişse onları ekle
        if 'title' in data:
            update_data['title'] = data['title']
        if 'description' in data:
            update_data['description'] = data['description']
        if 'status' in data:
            update_data['status'] = data['status']
            
        # Adımı güncelle
        step_ref.update(update_data)
        
        return jsonify({'message': 'Adım başarıyla güncellendi'}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Tüm görevlerin adımlarını getir (Admin/SuperAdmin için)
@app.route('/admin/tum-adimlar', methods=['GET'])
def get_all_task_steps():
    try:
        user_email = request.headers.get('X-User-Email')
        
        if not user_email:
            return jsonify({'error': 'Kullanıcı bilgisi gerekli'}), 401
            
        # Kullanıcı rolünü kontrol et
        user_ref = db.collection("users").where("email", "==", user_email).limit(1).get()
        if not user_ref:
            return jsonify({'error': 'Kullanıcı bulunamadı'}), 404
            
        user_data = user_ref[0].to_dict()
        user_role = user_data.get('role')
        
        if user_role not in ['admin', 'superAdmin']:
            return jsonify({'error': 'Bu işlem için yetkiniz yok'}), 403
            
        # Tüm adımları getir
        steps_ref = db.collection('task_steps').order_by('created_at', direction=firestore.Query.DESCENDING)
        steps = steps_ref.stream()
        
        # Görev bilgilerini de al
        tasks_ref = db.collection('tasks')
        tasks_dict = {}
        
        for task in tasks_ref.stream():
            task_data = task.to_dict()
            tasks_dict[task.id] = task_data
        
        steps_list = []
        for step in steps:
            step_data = step.to_dict()
            step_data['id'] = step.id
            step_data['status_tr'] = get_status_turkish(step_data.get('status', 'pending'))
            
            # İlgili görev bilgisini ekle
            task_id = step_data.get('task_id')
            if task_id in tasks_dict:
                step_data['task_info'] = tasks_dict[task_id]
            
            steps_list.append(step_data)
            
        return jsonify(steps_list), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Hata yakalayıcılar
@app.errorhandler(404)
def not_found(e):
    return jsonify({"error": "Endpoint bulunamadı"}), 404

@app.errorhandler(403)
def forbidden(e):
    return jsonify({"error": "Bu işlem için yetkiniz yok"}), 403

if __name__ == '__main__':
    app.run(debug=True, port=5001) 